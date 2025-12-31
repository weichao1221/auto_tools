import argparse
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill


def get_valid_excel_column(prompt_text):
    """
    循环输入，直到用户输入合法的 Excel 列标（A ~ XFD）
    自动转换为大写
    """
    while True:
        col = input(prompt_text).strip().upper()

        # 必须是字母
        if not col.isalpha():
            print("❌ 输入无效：只能输入字母，例如 A、B、AA、XFD")
            continue

        # Excel 最大列：XFD（16384 列）
        try:
            idx = column_index_from_string(col)
        except ValueError:
            print("❌ 输入无效：不是合法的 Excel 列标，请重新输入")
            continue

        # 合法
        return col
    
# -----------------------------
# 1. 命令行参数解析
# -----------------------------
parser = argparse.ArgumentParser(description="Excel IQR 异常值检测脚本")
parser.add_argument("--file", required=True, help="输入 Excel 文件路径")
parser.add_argument("--output", help="输出文件路径（可选）")
args = parser.parse_args()

input_file = args.file
output_file = args.output

if not output_file:
    output_file = input_file.replace(".xlsx", "") + "_result.xlsx"

print(f"输入文件：{input_file}")
print(f"输出文件：{output_file}")

# -----------------------------
# 2. 加载工作簿
# -----------------------------
wb = load_workbook(input_file, data_only=True)

print("选择要处理的工作表：")
for idx, name in enumerate(wb.sheetnames, start=1):
    print(f"{idx}: {name}")

sheet_index = int(input("请输入要处理的工作表序号："))
sheetname = wb.sheetnames[sheet_index - 1]
print(f"处理工作表：{sheetname}")

# -----------------------------
# 3. 用户输入参数
# -----------------------------
input_start_letter = get_valid_excel_column("请输入开始列字母：")
input_end_letter = get_valid_excel_column("请输入结束列字母：")
input_start_row_index = int(input("请输入开始行索引："))
input_multiplier = input("请输入IQR倍数(默认1.5，数字越大，不合格的越多)：")

if not input_multiplier:
    multiplier = 1.5
else:
    multiplier = float(input_multiplier)

print(f"开始列：{input_start_letter}，结束列：{input_end_letter}，IQR倍数：{multiplier}")

start_col_index = column_index_from_string(input_start_letter)
end_col_index = column_index_from_string(input_end_letter)

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

ws = wb[sheetname]

row_positions = []

# -----------------------------
# 4. 读取数据并保存 cell 对象
# -----------------------------
for row_idx, row in enumerate(
        ws.iter_rows(min_row=input_start_row_index,
                     min_col=start_col_index,
                     max_col=end_col_index),
        start=input_start_row_index):

    print(f"处理行：{row_idx}")

    numeric_values = []
    cell_objects = []

    for cell in row:
        cell_objects.append(cell)
        if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
            numeric_values.append(cell.value)

    if numeric_values:
        row_positions.append({
            "row": row_idx,
            "cells": cell_objects,
            "values": numeric_values
        })

# -----------------------------
# 5. IQR 异常值检测
# -----------------------------
outlier_count = 0

for row_info in row_positions:
    values = row_info["values"]

    if len(values) >= 4:  # 至少 4 个数据点才能计算 IQR
        q1 = np.percentile(values, 25)
        q3 = np.percentile(values, 75)
        iqr = q3 - q1

        lower_bound = q1 - multiplier * iqr
        upper_bound = q3 + multiplier * iqr

        for cell in row_info["cells"]:
            val = cell.value
            if isinstance(val, (int, float)):
                if val < lower_bound or val > upper_bound:
                    cell.fill = yellow_fill
                    outlier_count += 1

# -----------------------------
# 6. 保存文件
# -----------------------------
wb.save(output_file)

print(f"处理完成，共处理行数：{len(row_positions)}，异常数据单元格数量：{outlier_count}")
input("按回车键退出...")